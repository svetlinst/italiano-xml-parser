import datetime
import os
from pathlib import Path
import xmltodict
from models.erp_models import InvoiceParser, XmlToInvoiceMapper
import logging
import shutil
from openpyxl import load_workbook

# Config path to directories
project_dir = Path(os.path.dirname(os.path.abspath(__file__)))
input_dir = Path(project_dir / 'input')
templates_dir = Path(project_dir / 'templates')
output_dir = Path(project_dir / 'output')

# Constants
INVOICE_TEMPLATE_NAME = 'base_invoice_template.xlsx'
invoice_xl_template = templates_dir / INVOICE_TEMPLATE_NAME

# Configure the logger
LOG_NAME = 'log.csv'
logging.basicConfig(filename=LOG_NAME,
                    filemode='a',
                    format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                    datefmt='%H:%M:%S',
                    level=logging.DEBUG)
logging.info("Logging execution for XML-Parser")

# Check if there are XML files in the input directory
files_available = [x for x in input_dir.iterdir() if x.suffix == '.xml']

if len(files_available) == 0:
    print('No files found.')
    logging.info('No files found. Break the execution...')
    quit()

# Iterate over all XML input files
invoice_parser = None
for file in files_available:
    # Open read the XML file into a dictionary
    with open(file) as fd:
        xml_doc = xmltodict.parse(fd.read())

    # Create an instance of the InvoiceParser object with the dictionary data
    invoice_parser = InvoiceParser(xml_doc)

    # Try to parse the XML file into an Invoice object(s)
    try:
        invoice_parser.parse_invoice_data()
        logging.debug('Successfully parsed the file')
    except Exception as ex:
        logging.exception('message')
        raise ValueError('Error parsing the data!')

mapper = XmlToInvoiceMapper()

if invoice_parser:
    # Create a new directory that will contain the Excel output files
    output_sub_folder_name = 'invoices_' + datetime.datetime.now().strftime('%d-%m-%Y-%H-%M-%S')
    os.makedirs(output_dir / output_sub_folder_name)

    counter = 0
    for invoice in invoice_parser:
        counter += 1
        # Create a copy of the invoice template
        new_file_name = f'invoice_{counter}_{getattr(invoice.header, "FacturaID")}.xlsx'
        output = Path(output_dir / output_sub_folder_name) / new_file_name
        shutil.copyfile(invoice_xl_template, output)

        # Open the new copy and append the info from the invoice
        invoice_xl = load_workbook(output)
        ws = invoice_xl.worksheets[0]

        # ToDo: Apply formatting when updating the excel file
        # Update header
        for key, val in mapper.HEADER_FIELD_MAPPING.items():
            ws[val].value = getattr(invoice.header, key)

        # Update the line items
        line_items_cnt = len(invoice.detail.line_items)
        # Insert rows if needed
        if line_items_cnt > 1:
            # Offset the merged cells
            merged_cells_range = [x for x in ws.merged_cells.ranges if x.min_row > 23]
            for merged_cell in merged_cells_range:
                merged_cell.shift(0, line_items_cnt - 1)

            # Insert new rows where the line items will be written
            ws.insert_rows(25, line_items_cnt - 1)

        # Populate the data
        current_row = 23
        for li in invoice.detail.line_items:
            for key, val in mapper.LINE_ITEM.items():
                coord = val + str(current_row)
                ws[coord].value = getattr(li, key)
            current_row += 1

        # Save the workbook
        invoice_xl.save(output)
